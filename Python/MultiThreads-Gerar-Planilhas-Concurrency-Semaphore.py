import os
import pandas as pd
from openpyxl import load_workbook
import threading

#----(MOCKS)
type = 'Inventory'
chooseFile = r'C:\Users\sergio.silva\Downloads\Documentos\INVENTARIO_JUL_2023.xlsx'
output_dir = r'C:\Users\sergio.silva\Downloads\Documentos\Outputs'
model_path = r'C:\Users\sergio.silva\Downloads\Documentos\Modelos\modelo.xlsx'
data_path_contaclist = r'C:\Users\sergio.silva\Downloads\Documentos\CONTACT_LIST_2023.xlsx'
#----(MOCKS)

#----(AE)
# type = st[0].type
# output_dir = st[0].diretorioSaida
# model_path = st[0].caminhoPlanilhaModelo
# data_path_contaclist = st[0].caminhoPlanilhaContactList
#----(AE)

# Lock para sincronizar o acesso a dados_relatorio
data_lock = threading.Lock()
# Semaphore para limitar o número de threads
thread_limiter = threading.Semaphore(3)

dados_relatorio = [[], []]
        # {
        #     'Customer': '',
        #     'TipoDocumento': '',
        #     'DocumentoEnviado': '',
        #     'EmailEnviadoPara': '',
        #     'Status': '',
        #     'Descricao': ''
        # }

#---------------------------------------------------( METODOS )---------------------------------------------------

def create_excel_from_model_inventario(model_path, path_inventario, sheet_name_contact_list, sheet_name_inventario, path_contact_list, output_dir):
    # Ler a planilha de dados usando pandas
    df_inventario = pd.read_excel(path_inventario, sheet_name=sheet_name_inventario, engine='openpyxl', header=7)
    df_contact_list = pd.read_excel(path_contact_list, sheet_name=sheet_name_contact_list, engine='openpyxl', header=0)   

   # Verificar se cada valor da coluna G (índice 6) do df_inventario está na coluna A (índice 0) do df_contact_list
    for _, row in df_inventario.iterrows():
        nome_proveedor = row.iloc[6]

        # Verificar se o e-mail do fornecedor em df_inventario está presente em df_contact_list
        if pd.isna(nome_proveedor):
            continue
        elif nome_proveedor not in df_contact_list.iloc[:, 0].values:
            dados_relatorio.append({
                'Customer': nome_proveedor,
                'TipoDocumento': 'Inventário',
                'DocumentoEnviado': '',
                'EmailEnviadoPara': '',
                'Status': 'Erro',
                'Descricao': f"E-mail do fornecedor {nome_proveedor} inválido!"
            })
        else:
            matched_rows_with_X = df_contact_list[(df_contact_list.iloc[:, 0] == nome_proveedor) & (df_contact_list.iloc[:, 14].str.lower() == 'x')]
            values_from_column_G = matched_rows_with_X.iloc[:, 6].tolist()

            if not(len(values_from_column_G) >= 0):
                dados_relatorio.append({
                    'Customer': nome_proveedor,
                    'TipoDocumento': 'Inventário',
                    'DocumentoEnviado': '',
                    'EmailEnviadoPara': '',
                    'Status': 'Erro',
                    'Descricao': f"E-mail do fornecedor {nome_proveedor} sem marcação 'X' em 'Inventory'!"
                })
    
    threads = []

    for provider, group in df_inventario.groupby('PROVEEDOR'):
        # Adquire o semaphore (se já houver 3 threads em execução, isso irá bloquear até que uma delas termine)
        thread_limiter.acquire()
        
        thread = threading.Thread(target=process_provider, args=(provider, group, 'Inventory'))
        thread.start()
        threads.append(thread)

    # Aguarda todas as threads terminarem
    for thread in threads:
        thread.join()


def create_excel_from_model_consumo(model_path, path_consumo, sheet_name_contact_list, sheet_name_consumo, path_contact_list, output_dir):
    # Ler a planilha de dados usando pandas
    df_consumo = pd.read_excel(path_consumo, sheet_name=sheet_name_consumo, engine='openpyxl', header=2)
    df_contact_list = pd.read_excel(path_contact_list, sheet_name=sheet_name_contact_list, engine='openpyxl', header=0)   

   # Verificar se cada valor da coluna G (índice 4) do df_consumo está na coluna A (índice 0) do df_contact_list
    for _, row in df_consumo.iterrows():
        nome_customer = row.iloc[4]

        # Verificar se o e-mail do fornecedor em df_consumo está presente em df_contact_list
        if pd.isna(nome_customer):
            continue
        elif nome_customer not in df_contact_list.iloc[:, 0].values:
            dados_relatorio.append({
                'Customer': nome_customer,
                'TipoDocumento': 'Consumo',
                'DocumentoEnviado': '',
                'EmailEnviadoPara': '',
                'Status': 'Erro',
                'Descricao': f"E-mail do fornecedor {nome_customer} inválido!"
            })
        else:
            matched_rows_with_X = df_contact_list[(df_contact_list.iloc[:, 0] == nome_customer) & (df_contact_list.iloc[:, 13].str.lower() == 'x')]
            values_from_column_G = matched_rows_with_X.iloc[:, 6].tolist()

            if not(len(values_from_column_G) >= 0):
                dados_relatorio.append({
                    'Customer': nome_customer,
                    'TipoDocumento': 'Consumo',
                    'DocumentoEnviado': '',
                    'EmailEnviadoPara': '',
                    'Status': 'Erro',
                    'Descricao': f"E-mail do fornecedor {nome_customer} sem marcação 'X' em 'Consumo'!"
                })
    
    threads = []

    for provider, group in df_consumo.groupby('PROVEEDOR'):
        # Adquire o semaphore (se já houver 3 threads em execução, isso irá bloquear até que uma delas termine)
        thread_limiter.acquire()
        
        thread = threading.Thread(target=process_provider, args=(provider, group, 'Consumo'))
        thread.start()
        threads.append(thread)

    # Aguarda todas as threads terminarem
    for thread in threads:
        thread.join()

def process_provider(provider, group, tipo_documento):
    # Carregar a planilha modelo
    wb = load_workbook(model_path)
    ws = wb.active
    
    # Definir a linha de início para inserção como 9
    row_start = 9
    
    # Inserir os dados do provedor na planilha modelo
    for index, row in group.iterrows():
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_start, column=col_num, value=value)
        row_start += 1
    
    # Salvar a planilha modificada no diretório de saída
    wb.save(os.path.join(output_dir, f'{provider}.xlsx'))
    
    with data_lock:
        dados_relatorio.append({
            'Customer': provider,
            'TipoDocumento': tipo_documento,
            'DocumentoEnviado': '',
            'EmailEnviadoPara': '',
            'Status': 'Gerada',
            'Descricao': f"Planilha gerada com sucesso"
        })
    
    # Libera o semaphore
    thread_limiter.release()

##################################################( INVENTARIO )##################################################
if type == 'Inventory':
    output_dir = f'{output_dir}/Inventory'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    create_excel_from_model_inventario(model_path=model_path, 
                            path_inventario=chooseFile, 
                            sheet_name_inventario='Mes actual', 
                            sheet_name_contact_list='Contact List',
                            path_contact_list=data_path_contaclist,
                            output_dir=output_dir)
    
##################################################( CONSUMO )##################################################
if type == 'Consumption':
    output_dir = f'{output_dir}/Consumo'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    create_excel_from_model_consumo(model_path=model_path, 
                            path_consumo=chooseFile, 
                            sheet_name_consumo='CONSUMOS FOII.', 
                            sheet_name_contact_list='Contact List',
                            path_contact_list=data_path_contaclist,
                            output_dir=output_dir)

print(dados_relatorio)